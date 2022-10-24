import sys
import io
import paramiko
from apscheduler.schedulers.blocking import BlockingScheduler
from datetime import datetime, timedelta
from DBConnector import*
from SportsConverter import*
from SportsChecker import*
from collections import OrderedDict
from openpyxl import load_workbook
from SendMail import*
import traceback


sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

class SportsMonitor () :

    def __init__(self, config_path,date) :

        self.date = date

        self.config = configparser.ConfigParser()
        self.config.optionxform = str
        self.config.read(config_path,encoding='utf-8-sig')

        self.score_sql = self.config['sql']['score_sql']
        self.push_sql = self.config['sql']['push_days_sql']

        db_info = self.config['db']
        self.db_host = db_info['host']
        self.db_port = int(db_info['port'])
        self.db_user = db_info['user']
        self.db_pw = db_info['pw']
        self.def_db = db_info['db']

        self.sportsmapper = self.config['sport_ids']

        self.score_leagues = set(self.config['info']['score_leagues'].split(","))
        self.push_leagues = set(self.config['info']['push_leagues'].split(","))

        self.comple_path = self.config['info']['comple_path']
        self.n_comple_path = self.config['info']['n_comple_path']

        self.monitoring = OrderedDict() #key : id, value : game info
        self.nextmonitoring = OrderedDict() #key : id, value : game info

        self.score_db_results = OrderedDict() # key : id, value : game info
        self.push_db_results = OrderedDict() #key : id, value : game pushes
        self.web_results = OrderedDict()


    def record_monitoring_log(self) :
        date = self.date.strftime("%m%d")
        
        comple_log = configparser.ConfigParser()
        comple_log.optionxform = str
        comple_log.read(self.comple_path,encoding = 'utf-8-sig')

        n_comple_log = configparser.ConfigParser()
        n_comple_log.optionxform = str
        n_comple_log.read(self.n_comple_path,encoding = 'utf-8-sig')

        with open (self.comple_path, 'w+', encoding='UTF-8') as file :
            comple_log[date] = {}
            for id,game in self.monitoring.items() :
                comple_log.set(date,id,str(game))
            comple_log.write(file)

        with open (self.n_comple_path, 'w+', encoding='UTF-8') as file :
            n_comple_log[date] = {}
            for id,game in self.nextmonitoring.items() :
                tmp = ""
                for elem in game :
                    tmp = tmp + elem + ","
                tmp = tmp[:-1]

                n_comple_log.set(date,id,tmp)
            n_comple_log.write(file)
        
    def split_game_push(self, results) :

        game_push = []
        if len(results) == 0 :
            return

        b_id = results[0][PushElem.ID]
        id = -1

        for result in results :
            sport_id = result[PushElem.SPORT_ID]
            if self.sportsmapper.get(sport_id) is None :
                continue

            id = result[PushElem.ID]
            if id != b_id :
                if len(game_push) > 0 :                    
                    self.push_db_results[b_id] = game_push
                    game_push = []

            game_push.append(result)

            b_id = id
        self.push_db_results[b_id] = game_push

    
    def get_push_list_from_db(self) :

        connec = DBConnector(self.db_host,self.db_port,self.db_user,self.db_pw,self.def_db)
        yesterday = self.date - timedelta(days=1)

        sql = self.push_sql

        vars = [yesterday.strftime('%Y-%m-%d'), self.date.strftime('%Y-%m-%d')]

        results = connec.execute_sql(sql, vars)        

        self.split_game_push(results)

    def get_score_list_from_db(self) :
        
        today = self.date
        yesterday = today - timedelta(days = 1)
        connec = DBConnector(self.db_host,self.db_port,self.db_user,self.db_pw,self.def_db)
        vars = [yesterday.strftime('%Y-%m-%d'),today.strftime('%Y-%m-%d')]

        results = connec.execute_sql(self.score_sql, vars)
        for game in results :
            id = game[DBElem.ID]
            self.score_db_results[id] = game
        
    def adapt_score_db_results (self) :
        n_score_db_results = {}

        for id,game in self.score_db_results.items() :
            league = game[DBElem.LEAGUE]
            if id in self.monitoring and league in self.score_leagues:
                n_score_db_results[id] = game 

        n_score_db_results = OrderedDict( sorted(n_score_db_results.items(), key=lambda x : x[1][DBElem.TIME] ))
        self.score_db_results = n_score_db_results

    def adapt_push_db_results(self) :
        n_push_db_results = {}
        for id,pushes in self.push_db_results.items() :
            league = pushes[-1][PushElem.LEAGUE]
            if id in self.monitoring and league in self.push_leagues:
                n_push_db_results[id] = pushes 

        n_push_db_results = OrderedDict(sorted(n_push_db_results.items(), key=lambda x : x[1][-1][PushElem.START_TIME]))
        self.push_db_results = n_push_db_results
        
    def decide_monitoring_from_push_db(self) :

        monitor_day = self.date.strftime("%Y%m%d")
        for id,game_push in self.push_db_results.items() :

            starttime = datetime.strptime(game_push[-1][PushElem.START_TIME],"%Y%m%d%H%M").strftime("%Y%m%d")
            endtime = datetime.strptime(game_push[-1][PushElem.LOG_TIME], "%Y%m%d%H%M").strftime("%Y%m%d")
            end_status = game_push[-1][PushElem.STATUS]
            league = game_push[-1][PushElem.LEAGUE]
            
            if not (league in self.push_leagues or league in self.score_leagues) :
                continue

            if starttime < monitor_day :
                if endtime <monitor_day and (end_status == 'COMPLETED' or end_status == 'STOP_PLAYING'):
                    continue
                else :
                    self.monitoring[id] =game_push[-1]

            if starttime >= monitor_day and end_status!= 'COMPLETED' and end_status !='STOP_PLAYING' :
                self.nextmonitoring[id] = game_push[-1]

            if starttime == monitor_day and endtime ==monitor_day and (end_status =='COMPLETED' or end_status == 'STOP_PLAYING'):
                self.monitoring[id] = game_push[-1]
 

    def get_monitoring_from_db(self) :    

        self.get_push_list_from_db()
        self.get_score_list_from_db()

        # decide monitoring list
        self.decide_monitoring_from_push_db()
        self.record_monitoring_log()

        self.adapt_score_db_results()
        """
        print('score monitoring list')
        for id,game in self.score_db_results.items() :
            print(game) """

        self.adapt_push_db_results()
        """
        print('push monitoring list')
        for id,games in self.push_db_results.items() :
            print(games[-1])"""

    def get_web_results (self, path) :
        wb = load_workbook(path, data_only = True)

        for league in wb.sheetnames: 
            sheet = wb[league] 

            games = []

            for row_data in sheet.iter_rows(min_row=1): 
                game = []
                for cell in row_data: 
                    if cell.value is not None :
                        game.append(cell.value)

                if len (game) > 0 :
                    games.append(game)

            if len (games) > 0 : 
                self.web_results[league] = games
                    
        wb.close()


    def change_format (self) :

        converter = SportsConverter()
        self.web_results = converter.change_web_format(self.web_results)
        self.score_db_results = converter.change_db_format(self.score_db_results)
        self.push_db_results = converter.change_push_db_format(self.push_db_results)
    
def get_file_from_server(date) : 

    host = 'ec2-15-165-38-96.ap-northeast-2.compute.amazonaws.com'
    username = 'ubuntu'
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(host, username=username, key_filename='./key/20220321_.pem')
    sftp = ssh.open_sftp()
    sftp.get('/home/ubuntu/SPORTSALERT_MONITORING_CRAWLING_UBUNTU/CrawlingDatas/' + date +'.xlsx', './'+date+'.xlsx') 
    sftp.close()
    ssh.close()


def monitoring () :

    try : 
            now = datetime.now()

            #단건 실행 시 날짜설정  
            #db_date = datetime.strptime("2022-03-29","%Y-%m-%d") 
            db_date = now - timedelta(days = 1)
    
            get_file_from_server(now.strftime("%Y-%m-%d"))
            #get_file_from_server("2022-03-28")

            config_path = './Configuration/config.ini'    
            sm = SportsMonitor(config_path,db_date)
            sm.get_monitoring_from_db()
    
            sm.get_web_results('./' +  now.strftime("%Y-%m-%d") +  '.xlsx')
            #sm.get_web_results('./' + "2022-03-28" + '.xlsx')

            sm.change_format()
            checker = SportsChecker(sm.date, sm.monitoring, sm.web_results, sm.score_db_results, sm.push_db_results)
            checker.check_results()

    except :
            SendMail.send_simple_mail('에러 로그 메일',traceback.format_exc())
            print(traceback.format_exc())


# scheduler로 실행
"""sched = BlockingScheduler()
sched.add_job(monitoring, 'cron', hour = '12',minute="00", id="schdeuler_test")
sched.start()"""

#단건 실행

monitoring()



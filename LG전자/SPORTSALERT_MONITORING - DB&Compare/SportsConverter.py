from openpyxl import load_workbook
from time import gmtime, strftime
from Elem import*
import numpy as np
import configparser
import re

class SportsConverter :

    team_matching_path = "./Configuration/team_matching.xlsx"
    team_matching_push_path =  "./Configuration/team_matching_push.xlsx"
    config_path = "./Configuration/config.ini"
    team_dics = None
    push_team_dics = None
    sport_mapper = None
    league_mapper1 = None #from db to push
    league_mapper2 = None #from push to db

    # dic : key(team name in crawling site) value(team name in db)
    @staticmethod
    def make_team_matching_dic (sheet) :
        dic = {}
        for row in sheet.iter_rows(min_row = 2, max_col =3, min_col=1) :
            if row[0].value is None :
                break
            if row[1].value is not None :
                key = row[1].value
                value = row[2].value
                dic[key] = value

        return dic

    # dics : key(league name), value (dictionary of teamnames)
    @staticmethod
    def make_all_team_matching_dics(path) :
        workbook = load_workbook(path)
        dics = {}
        for sheet in workbook :
            dic = SportsConverter.make_team_matching_dic(sheet)
            dics[sheet.title] = dic

        return dics

    @staticmethod
    def make_sport_mapper_dic() :
        config = configparser.ConfigParser()
        config.optionxform = str
        config.read(SportsConverter.config_path,encoding='utf-8-sig')

        sport_mapper = {}
        for league in config['sports'] :
            sport = config['sports'][league]
            sport_mapper[league] = sport

        return sport_mapper

    
    @staticmethod
    def make_league_mapper_dic() :
        config = configparser.ConfigParser()
        config.optionxform = str
        config.read(SportsConverter.config_path,encoding='utf-8-sig')

        league_mapper1 = {}
        league_mapper2 = {}
        for db_league in config['league_mapper'] :
            push_league = config['league_mapper'][db_league]
            league_mapper1[db_league] = push_league
            league_mapper2[push_league] = db_league

        return league_mapper1,league_mapper2


    def __init__(self) :
        if SportsConverter.team_dics is None :
            SportsConverter.team_dics = SportsConverter.make_all_team_matching_dics(SportsConverter.team_matching_path)

        if SportsConverter.push_team_dics is None :
            SportsConverter.push_team_dics = SportsConverter.make_all_team_matching_dics(SportsConverter.team_matching_push_path)

        if SportsConverter.sport_mapper is None :
            SportsConverter.sport_mapper = SportsConverter.make_sport_mapper_dic()

        if SportsConverter.league_mapper1 is None :
            SportsConverter.league_mapper1,SportsConverter.league_mapper2 = SportsConverter.make_league_mapper_dic()
            
    def change_web_time_format(self,times) :

        n_times = []
        if str(type(times)) == "<class 'str'>" :
            time = times
            times = []
            times.append(time)

        for time in times :
            splited = time.replace(" ","").replace(":","").split(".")
            splited[2] = splited[2][0:4]   
            n_time = splited[1]+splited[0]+splited[2]

            year = strftime("%Y",gmtime())
            now = strftime("%m%d%H%M", gmtime()) 
            if(n_time > now) :
                n_time = str(int(year)-1) + n_time
            else :
                n_time = year + n_time
            n_times.append(n_time)
        
        """if len(n_times) == 1 :
            return n_times[0]"""
        return n_times
    
    def change_web_team_format(self,league,team_names) :

        n_teams = []

        dics = SportsConverter.team_dics
        for team in team_names : 
            if dics.get(league) is not None and dics.get(league).get(team):
                n_name = dics.get(league).get(team)
            else :
                n_name = team
            n_teams.append(n_name)
        
        return n_teams

    def change_web_score_format(self,league,scores) :

        n_scores = scores
        if league == 'IPL' :
            n_scores = []
            for score in scores :
                splited = score.split('/')
                n_score = splited[0]
                n_scores.append(n_score)

        if league == 'AFL' :
            n_scores = []
            for score in scores :
                if "(" in score :
                    n_score = re.sub('[(][.|0-9]+[)]','',score)
                    n_scores.append(int(n_score))
        return n_scores

    def change_web_push_format(self, league, pushes) :

        sport = self.sport_mapper.get(league)
        results = []

        for num,push in enumerate(pushes) :
            print(push) 
            push = push.split(",")
            result = []
            """if sport == 'FOOTBALL' :
                result = [[0]*2,[0]*2]

                for i,score in enumerate(push) :
                    print(score)
                    if score == "" :
                        result[0][i] = ""
                        result[1][i] = ""
                    else :
                        result[0][i] = int(score[1:-1].split("-")[0])
                        result[1][i] = int(score[1:-1].split("-")[1])"""

            if sport == 'RUGBY LEAGUE' or sport == 'BASKETBALL' or sport == 'AMERICAN FOOTBALL' or sport == 'ICEHOCKEY' or sport == 'BASEBALL' or sport == 'CRICKET' or sport == 'FOOTBALL' :
            
                h_score = 0
                a_score = 0
                result = [[0]*int(len(push)/2), [0]*int(len(push)/2)]
                
                print(push)
                for i,score in enumerate(push) :
                    
                    if score == "" or score == "X" or score == 'x':
                        if i% 2 == 0 :
                            result[0][int(i/2)] = ''
                        else :
                            result[1][int(i/2)] = ''
                        continue

                    # home score
                    if i %2 == 0 :
                        h_score = h_score + int(score)
                        result[0][int(i/2)] = h_score

                    # away score
                    else :
                        a_score = a_score + int(score)
                        result[1][int(i/2)] = a_score


                if (sport == 'ICEHOCKEY') and result[0][4] != '' and result[0][1] != '':
                    if result[0][4] > result[1][4] :
                        result[0][4] = result[0][3] + 1
                        result[1][4] = result[1][3]
                    else :
                        result[1][4] = result[1][3] +1
                        result[0][4] = result[0][3]
            results.append(result)
        
        return results

    def change_web_format(self, web_results) :
        n_web_results = {}

        for league, game_list in web_results.items() :

            game_list = np.array(game_list).T.tolist()
            
            if len(game_list) == 0:
                n_web_results[league] = []
                continue

            times = game_list[WEBElem.TIME]
            homes = game_list[WEBElem.HOME]
            aways = game_list[WEBElem.AWAY]
            h_scores = game_list[WEBElem.HOME_SCORE]
            a_scores = game_list[WEBElem.AWAY_SCORE]

            game_list[WEBElem.TIME] = self.change_web_time_format(times)
            game_list[WEBElem.HOME] = self.change_web_team_format(league,homes)
            game_list[WEBElem.AWAY] = self.change_web_team_format(league,aways)
            game_list[WEBElem.HOME_SCORE] = self.change_web_score_format(league,h_scores)
            game_list[WEBElem.AWAY_SCORE] = self.change_web_score_format(league,a_scores)

            if len (game_list)-1 >= int(WEBElem.PUSH)  :
                pushes = game_list[WEBElem.PUSH]
                game_list[WEBElem.PUSH] = self.change_web_push_format(league, pushes)

            n_web_results[league] = np.array(game_list).T.tolist()

        return n_web_results

    def change_db_score_foramt(self, sport, score) :
        if sport == 'CRICKET' :
            splited = score.split('/')
            score = splited[0]
            
        return score

    def change_db_time_format(self,time) :
        n_time = time.replace("-","").replace(" ","").replace(":","")
        return n_time
         
    def change_db_format(self, db_results) :
            
        n_db_results = {}

        for  game in db_results.values() :

            game = list(game)
            
            league = game[DBElem.LEAGUE]
            h_score = game[DBElem.HOME_SCORE]
            a_score = game[DBElem.AWAY_SCORE]
            time = game[DBElem.TIME]

            sport = self.sport_mapper.get(league)

            game[DBElem.HOME_SCORE] = self.change_db_score_foramt(sport,h_score)
            game[DBElem.AWAY_SCORE] = self.change_db_score_foramt(sport,a_score)
            game[DBElem.TIME] = self.change_db_time_format(time)
            if league in SportsConverter.league_mapper1.keys() :
                league = SportsConverter.league_mapper1.get(league)
            
            if not(league in n_db_results) :
                n_db_results[league] = []
            n_db_results[league].append(game)
        
        return n_db_results


    def change_push_team_format(self,league,team) :

        n_name = None
        dics = SportsConverter.push_team_dics
        if dics.get(league) is not None  :
            if dics.get(league).get(team):
                n_name = dics.get(league).get(team)
            
        if n_name is None :
            n_name = team
        
        return n_name

    # dic 만들기 : key id, value: changed push list
    # id,pushlist 
    def change_push_db_format(self,push_db_results) :

        n_push_db_results = {}

        for id,pushes in push_db_results.items() :
            
            league = pushes[-1][PushElem.LEAGUE]
            sport = self.sport_mapper.get(league)
            s_time = pushes[-1][PushElem.START_TIME]
            home = self.change_push_team_format(league, pushes[-1][PushElem.HOME])
            away =self.change_push_team_format(league, pushes[-1][PushElem.AWAY])

            changed = [league,s_time, home, away]
            result = None

            if sport == 'BASEBALL' :
                result = [['']*10, ['']*10]
                pre_h_score = 0
                pre_a_score = 0
                pre_state_num = 1

                for push in pushes :
                    state = push[PushElem.STATE]
                    h_score = push[PushElem.SCORES].split(":")[1]
                    a_score = push[PushElem.SCORES].split(":")[0]
                    state_num = int(state[0:-1])

                    if state_num >=10 :
                        result[0][9] = h_score
                        result[1][9] = a_score
                        continue
                    
                    result[0][state_num-1] = h_score
                    result[1][state_num-1] = a_score

                    if state_num - pre_state_num >=2  :
                        for i in range(pre_state_num, state_num-1) :
                            result[0][i] = pre_h_score
                            result[1][i] = pre_a_score
                   
                    if push[PushElem.STATUS] == 'COMPLETED':
                        last_state = state[-1]
                        if last_state == '▲' :
                            for i in range(pre_state_num,state_num-1):
                                result[0][i] = pre_h_score 
                                result[1][i] = pre_a_score 
                            result[0][state_num-1] = ''
                            result[1][state_num-1] = pre_a_score

                        else :
                            for i in range(pre_state_num,state_num):
                                result[0][i] = pre_h_score 
                                result[1][i] = pre_a_score 
                            

                    pre_h_score = h_score
                    pre_a_score = a_score
                    pre_state_num = state_num
                    
                               
            if sport == 'AMERICAN FOOTBALL' or sport == 'BASKETBALL':
                result = [['']*5, ['']*5] 

                pre_state_num = 0
                pre_h_score = 0
                pre_a_score = 0

                for push in pushes :
                    # 연장전
                    if "OT" in push[PushElem.STATE] :
                        continue

                    state_num = int(push[PushElem.STATE][0])
                    h_score = push[PushElem.SCORES].split(":")[1]
                    a_score = push[PushElem.SCORES].split(":")[0]

                    result[0][state_num-1] = h_score
                    result[1][state_num-1] = a_score

                    if state_num - pre_state_num >= 2 :
                        for num in range(pre_state_num, state_num-1) :
                            result[0][num] = pre_h_score
                            result[1][num] = pre_a_score
                    pre_state_num = state_num
                    pre_h_score = h_score
                    pre_a_score = a_score
                
                last_push = pushes[-1]
                if "OT" in last_push[PushElem.STATE] :
                    h_score = last_push[PushElem.SCORES].split(":")[1]
                    a_score = last_push[PushElem.SCORES].split(":")[0]
                    result[0][-1] = h_score
                    result[1][-1] = a_score

            if sport == 'FOOTBALL' :
                result = [['']*2, ['']*2]
                pre_state_num = 1
                pre_h_score = 0
                pre_a_score = 0

                for push in pushes :

                    h_score = push[PushElem.SCORES].split(":")[1]
                    a_score = push[PushElem.SCORES].split(":")[0]
                    pso_check = False

                    if (push[PushElem.STATE] == 'ET') :
                        continue


                    elif(push[PushElem.STATE] == 'PSO') :
                        if(not pso_check) :
                            result[0][pre_state_num-1] = pre_h_score
                            result[1][pre_state_num-1] = pre_a_score
                        continue

                
                    else :
                        state_num = int(push[PushElem.STATE][0])

                        if state_num != pre_state_num :
                            result[0][pre_state_num-1] = pre_h_score
                            result[1][pre_state_num-1] = pre_a_score

                        pre_state_num = state_num

                        pre_h_score= h_score
                        pre_a_score = a_score

            if sport == 'RUGBY LEAGUE' :
                result = [['']*3, ['']*3] 

                state_dic = { '1HALF' : 0, '2HALF' :1, '1OT' :2}

                for i,push in enumerate(pushes) :
                    h_score = push[PushElem.SCORES].split(":")[1]
                    a_score = push[PushElem.SCORES].split(":")[0]

                    state = push[PushElem.STATE]
                    idx = state_dic.get(state)

                    result[0][idx] = h_score
                    result[1][idx] = a_score
            
            if sport == 'ICEHOCKEY' :
                result = [['']*5, ['']*5] 

                state_dic = { '1P' : 0, '2P' :1, '3P' :2, '1OT': 3,'2OT':3, '3OT':3, 'SO' :4}

                for i,push in enumerate(pushes) :
                    h_score = push[PushElem.SCORES].split(":")[1]
                    a_score = push[PushElem.SCORES].split(":")[0]

                    state = push[PushElem.STATE]
                    idx = state_dic.get(state)

                    result[0][idx] = h_score
                    result[1][idx] = a_score
                    
                for i in range (0,3) :
                    if result[0][i] == '' :
                        result[0][i] = result[0][i-1]
                    if result[1][i] == '' :
                        result[1][i] = result[1][i-1]


                if pushes[-1][PushElem.STATE] == 'SO' :
                    if result[0][3] == '' : 
                        result[0][3] = result[0][2]
                    if result[1][3] == '' :
                        result[1][3] = result[1][2]

            if result is not None :
                changed.append(result)
                n_push_db_results[id] = changed
        
        return n_push_db_results
"""
web_results = {}
web_results ['MLB'] = [['01.11. 18:00', 'Miami Dolphins', 'Los Angeles Rams', '28', '17', ['7', '7', '21', '3', '0', '0', '0', '7', '', '']]]
print(web_results)
converter = SportsConverter()
web_results = converter.change_web_format(web_results)
"""

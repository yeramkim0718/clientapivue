import requests
import sys
import io
import configparser
import json
import time
from openpyxl import load_workbook
from Elem import*
from datetime import datetime


sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

class SportsWebOS5 :  

    def make_sports_data(self, excel_path) :
        self.sports = {}
        self.leagues = {}
        self.teams = {}

        wb = load_workbook(excel_path, data_only = True)
  
        wses = ['sport','league','team']

        for i,ws in enumerate(wses) :
            ws = wb[ws]
            res = {}
            for row in range(2, 10000) :
                first = ws.cell(row,1).value
                if first is None :
                    break
                one = []
                key = None
                for col in range(1, 10000) :
                    val = ws.cell(row,col).value
                    if val is None  :
                        break
                    if col == 1 :
                        key = val
                    one.append(val)
                res[key] = one

            if i == 0 :
                self.sports = res
            elif i == 1 :
                self.leagues = res
            elif i == 2:
                self.teams = res

            
    def __init__ (self, config_path, excel_path) :
        self.config = configparser.ConfigParser()
        self.config.optionxform = str
        self.config.read(config_path,encoding='utf-8-sig')

        self.headers = {}
        for key, val in self.config['HEADERS'].items() :
            self.headers[key] = val
        
        self.params = {}
        for key,val in self.config['PARAMS'].items() :
            self.params[key] = val

        self.url_initService = self.config['URL']['initService']
        self.url_getSportList = self.config['URL']['getSportList']
        self.url_getSportProgramList = self.config['URL']['getSportProgramList']

        self.make_sports_data(excel_path)
        self.err_log = open('./LOG/'+datetime.utcnow().strftime("%Y%m%d%H%M%S")+".txt",'wt',encoding='UTF8')
        

    def initAuthentication (self) :

        response = requests.post(self.url_initService, headers=self.headers, params = self.params)
        dict = json.loads(response.text)

        authen = dict.get('initServices').get('authentication').get('sessionID')
        self.headers['X-Authentication'] = authen
        
        print("init authenticiation : "+ authen)

    def getLangnum(self, lang) :

        if lang == 'ko-KR' :
            return Lang.ko_KR
        if lang == 'ja-JP' :
            return Lang.ja_JP
        if lang == 'fr-FR' :
            return Lang.fr_FR
        if lang == 'es-ES' :
            return Lang.es_ES
        if lang == 'en-US' :
            return Lang.en_US
        if lang == 'en-GB' :
            return Lang.en_GB
        if lang == 'de-DE' :
            return Lang.de_DE
        if lang == 'nl-NL' :
            return Lang.nl_NL

    def changeLang(self,lang) :
        self.headers['X-Device-Language'] = lang

    def changeServer(self, server) :
        print("changeserver: "+server)
        self.err_log.write(server+"\n")

        self.url_getSportList = "http://" + server + self.config['URL']['getSportList']
        self.url_getSportProgramList = "http://" + server + self.config['URL']['getSportProgramList']

        if server == 'kic' :
            self.url_initService = 'http://KR'+self.config['URL']['initService']
        elif server == 'eic' :
            self.url_initService = 'http://GB'+self.config['URL']['initService']
        elif server == 'aic' :
            self.url_initService = 'http://US'+self.config['URL']['initService']

        self.initAuthentication()

    def reqSportList (self, request_type, id) :
        data = json.dumps({'request_type' : request_type, 'id':id},indent = 3)
        try :
            response = requests.request("POST",self.url_getSportList, headers = self.headers, data = data,timeout=1)
            result = json.loads(response.text)
            return result
        except requests.exceptions.Timeout:
            self.err_log.write("time_out\n")
            return self.reqSportList(request_type,id)
        except requests.exceptions.ConnectionError :
            self.err_log.write("connection_error : 같은 IP로부터 짧은시간에 너무 많은 요청을 보냄.\n")
            time.sleep(0.1)
            return self.reqSportList(request_type,id)

# server, language 추가 
    def getSportList (self,lang) : 

        api_sport = {}
        api_league = {}
        api_team = {}

        self.changeLang(lang)

        # 1. sportList
        res = self.reqSportList('1','')
        if res is not None : 
            for re in res.get('list') :
                id = re.get('id')
                li = [re.get('logo'), re.get('bg_img'),re.get('name')]
                api_sport[id] = li

        # 2. LeagueList
        for sport_id in api_sport.keys() :
            res = self.reqSportList('2',sport_id)
            if res is not None : 
                for re in res.get('list') :
                    id = re.get('id')
                    li = [re.get('logo'),sport_id,re.get('name')]
                    api_league[id] = li
        
        # 3. TeamList
        for league_id in api_league.keys() :
            res = self.reqSportList('3',league_id) 
            if res is not None : 
                for re in res.get('list'):
                    id = re.get('id')
                    li = [league_id,re.get('name')]
                    api_team[id] = li

        return api_sport,api_league,api_team

    def convert_for_compare(self, lang) :
        self.c_sports = {}
        self.c_leagues = {}
        self.c_teams = {}

        n_lang = self.getLangnum(lang)
        for id, sport in self.sports.items() :
            logo = sport[Sport.logo]
            bg_img = sport[Sport.bg_img]
            name = sport[n_lang]
            self.c_sports[id] = [logo,bg_img,name]

        for id, league in self.leagues.items() :
            logo =league[League.logo]
            sport_id = league[League.sport_id]
            name = league[n_lang]
            self.c_leagues[id] = [logo,sport_id,name]
        
        for id, team in self.teams.items() :
            league_id = team[Team.league_id]
            name = team[n_lang]
            self.c_teams[id] = [league_id,name]

    def check_u_getSportList(self, api, comp) :
        for id, sport in api.items() :
            c_sport = comp.get(id)
            if c_sport is None :
                print("잘못된 ID가 포함되 있습니다.")
                self.err_log.write("잘못된 스포츠 ID가 포함되 있습니다. " + id +"\n")
                continue
            for i,e in enumerate(sport) :
                if(c_sport[i]) != e :
                    print("잘못된 값이 응답에 왔습니다." + str(i))
                    print(sport)
                    print(c_sport)
                    self.err_log.write("잘못된 값이 응답에 왔습니다. :" + id+"\n")
                    self.err_log.write("응답 : " + str(sport)+"\n")
                    self.err_log.write("원래값 : " + str(c_sport) +"\n")
            del comp[id]

        if len(comp) > 0 :
            print("응답에 오지 않은 값이 있습니다.")
            print(comp)        
            self.err_log.write("응답에 오지 않은 값이 있습니다.\n")
            for id,c_sport in comp.items() :
                self.err_log.write(id + " : " + str(c_sport)+"\n")

    def check_getSportList (self, lang) :
 
        api_sport,api_league,api_team = self.getSportList(lang)
        self.convert_for_compare(lang) 

        self.check_u_getSportList(api_sport, self.c_sports)
        self.check_u_getSportList(api_league, self.c_leagues)
        self.check_u_getSportList(api_team, self.c_teams)

    def reqSportProgramList (self, id) :
        data = json.dumps({'id' : id, 'device_src_idx':'7','chan_code_list':''},indent = 3)
        try : 
            response = requests.request("POST",self.url_getSportProgramList, headers = self.headers, data = data,timeout=1)
            result = json.loads(response.text)
            return result
        except requests.exceptions.Timeout:
            self.err_log.write("time_out\n")
            return self.reqSportProgramList(id)
        except requests.exceptions.ConnectionError :
            self.err_log.write("connection_error : 같은 IP로부터 짧은시간에 너무 많은 요청을 보냄.\n")
            time.sleep(0.1)
            return self.reqSportProgramList(id)


    def getSportProgramList(self, lang) :
        api_program= {}
        self.changeLang(lang)

        for team_id in self.teams.keys() :
            res = self.reqSportProgramList(team_id)
            if res is not None : 
                list = res.get('list')
                api_program[team_id] = list
                

        return api_program
    
    def check_u_getSportProgramList(self,id, res,lang) :

        for game in res :
            #1. sportname, sport_id, league_id, league_name 값이 team_id에 맞는 값인가?
            c_league_id = self.teams.get(id)[Team.league_id]
            c_league_name = self.leagues.get(c_league_id)[self.getLangnum(lang)]
            c_sport_id = self.leagues.get(c_league_id)[League.sport_id]
            c_sport_name = self.sports.get(c_sport_id)[self.getLangnum(lang)]
            comp = [c_sport_name, c_sport_id, c_league_id, c_league_name]

            target = [game.get('sport_name'),game.get('sport_id'),game.get('league_id'),game.get('league_name')]
            for i,t in enumerate(target) :
                if t != comp[i] :
                    print("잘못된 값이 응답에 왔습니다." + id)
                    print(target)
                    print(comp)
                    self.err_log.write("잘못된 값이 응답에 왔습니다. team_id : " + id +"\n")
                    self.err_log.write("응답 : " + str(target)+"\n")
                    self.err_log.write("원래와야하는 값 : " + str(comp) + "\n")

            
            # 2. participant 모두 각 id에 따른 name이 맞는가? 
            parties = game.get('participant_list')
            ids = []
            for part in parties :
                t_id = part.get('id')
                ids.append(t_id)
                name = part.get('name')
                c_name = self.teams.get(t_id)[self.getLangnum(lang)]
                if name != c_name :
                    print("잘못된 값이 응답에 왔습니다." )
                    print(t_id + " : " + name)
                    print(id + " : " +c_name)
                    self.err_log.write("team id와 name이 맞지 않습니다. \n")
                    self.err_log.write("응답 값 : " + t_id + " " + name +"\n")
                    self.err_log.write("원래 값 : " + c_name + "\n")


            #3. 검색한 id에 맞는 team의 경기인가?
            if not (id in ids) :
                print("검색한 id에 맞지 않는 경기가 왔습니다." + id)
                print(ids)
                self.err_log.write("검색한 id에 맞지 않는 경기가 왔습니다. : " + id + "\n")
                break
            #4. game_id값이 일관되는가?
            g_id1 = game.get('game_id')
            g_id2 = game.get('participant_list')[0].get('game_id')
            g_id3 = game.get('participant_list')[1].get('game_id')

            if not (g_id1 == g_id2 and g_id2 == g_id3) : 
                print("game_id값이 일관되지 않습니다." + id)
                print(game)
                self.err_log.write("game_id값이 일관되지 않습니다." + id + "\n")
                self.err_log.write(g_id1 + " " + g_id2 + " " + g_id3 +"\n")

            #5. is_first_pos값이 모순되지 않는가?
            pos1 = game.get('participant_list')[0].get('is_first_pos')
            pos2 = game.get('participant_list')[1].get('is_first_pos')
            
            if not((pos1 == True and pos2 == False) or (pos1 == False and pos2 ==True)) :
                print('is_first_pos값이 잘못되었습니다. : ' +id)
                self.err_log.write('is_first_pos값이 잘못되었습니다. : ' + id +"\n")
                self.err_log.write(str(game)+"\n")
                print(str(pos1) + " " + str(pos2))
            #6 locaition값이 모순되지 않는가?
            loc1 = game.get('participant_list')[0].get('location')
            loc2 = game.get('participant_list')[1].get('location')
            
            if not((loc1 == 'HOME' and loc2 == 'AWAY') or (loc1 == 'AWAY' and loc2 =='HOME')) :
                print('location값이 잘못되었습니다. : ' + id)
                print(loc1 + " " + loc2)
                self.err_log.write('location값이 잘못되었습니다. : ' + id +"\n")
                self.err_log.write(loc1 + " " + loc2+"\n")

            #7 과거의 game이왔는가?
            endtime = game.get('game_end_time')
            if endtime != "" :
                now = datetime.utcnow().strftime("%Y%m%d%H%M%S")
                if endtime < now :
                    print("지난 경기가 응답에 왔습니다. : " +id)
                    print(endtime)
                    self.err_log.write("지난경기가 응답에 왔습니다. :" + id + "\n")
                    self.err_log.write("endtime : " + endtime + " now : " + now +" \n")

            
    def check_getSportProgramList(self,lang) :

        api_program = self.getSportProgramList(lang)

        for id,res in api_program.items() :
            self.check_u_getSportProgramList(id,res,lang)


sport = SportsWebOS5('./Configuration/config.ini','./Configuration/SportsAlert_5.0.xlsx')
langs = ['ko-KR','ja-JP','fr-FR','es-ES','en-US','en-GB','de-DE','nl-NL']

servers = ['aic','eic','kic']

for server in servers :
    sport.changeServer(server)
    for lang in langs : 
        print(lang)
        sport.err_log.write(lang +"\n")
        sport.err_log.write('getSportList\n')
        sport.check_getSportList(lang)
        sport.err_log.write('getSportProgramList\n')
        sport.check_getSportProgramList(lang)




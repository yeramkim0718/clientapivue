import sys
import io
from apscheduler.schedulers.blocking import BlockingScheduler 
from SeleniumCrawler import*
import configparser
from Elem import *
from ordered_set import OrderedSet
from collections import OrderedDict
import openpyxl
from datetime import date
import re
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

class SportsWebCrawler (SeleniumCrawler) :

    def make_crawle_info(self,config_path) :

        self.config = configparser.ConfigParser()
        self.config.optionxform = str
        self.config.read(config_path,encoding='utf-8-sig')

        self.detail_url = self.config['info']['detail_url']
        self.more_tag = self.config['info']['more_tag']
        self.prefs = self.config._sections.get('prefs')
        self.arguments = [*self.config._sections.get('argument')]

        self.crawle_leagues =  set()
        for lPerSport in self.config['leagues'].values() :
            self.crawle_leagues.update(lPerSport.split(","))
        
        for league in self.crawle_leagues :

            self.crawle_info[league] = [-1]*len(CRAWLEINFO)

            total_tags= []
            score_tags = []
            push_tags = []
            
            url = self.config['url'][league]

            if league in self.config['total_tag'] :
                total_tags = self.config['total_tag'][league]

            if league in self.config['score_tags'] :
                score_tags = self.config['score_tags'][league].split(",")

            if league in self.config['push_tags'] :
                push_tags = self.config['push_tags'][league].split(",")

            self.crawle_info[league][CRAWLEINFO.URL] = url
            self.crawle_info[league][CRAWLEINFO.TOTALTAG] = total_tags
            self.crawle_info[league][CRAWLEINFO.SCORETAG] = score_tags
            self.crawle_info[league][CRAWLEINFO.PUSHTAG] = push_tags

    def __init__(self, browser,config_path) :
        SeleniumCrawler.__init__(self,browser)

        self.config = None
        self.more_tag = None
        self.prefs = None
        self.arguments = None

        self.crawle_leagues = OrderedSet()
        self.crawle_info = OrderedDict() # key : league, value : url, list of tags (0 : url, 1 : total_tags, 2: score_tags, 3: push_tags)

        self.web_results = OrderedDict() # key : league, value : game info

        self.make_crawle_info(config_path)

    # setting to gmt time+0
    def set_gmt_zero (self) :  
        try :          
            self.driver.find_element_by_css_selector('.header__button.header__button--settings').click()
            self.driver.find_element_by_css_selector('#livescore-settings-form > div:nth-child(2) > div > div.timeZoneLine__icon').click()
            self.driver.find_element_by_xpath('//*[@class="timeZone__link"][contains(.,"GMT+0")]').click()
            self.driver.implicitly_wait(1)
            self.driver.find_element_by_xpath('//*[@id="lsid-window-close"]').click()
        except StaleElementReferenceException as e :
            print(e.msg)
            self.driver.find_element_by_xpath('//*[@class="timeZone__link"][contains(.,"GMT+0")]').click()
            self.driver.implicitly_wait(1)
            self.driver.find_element_by_xpath('//*[@id="lsid-window-close"]').click()

    def make_web_game_list (self,soup, total_tag, score_tags, push_tags) :
        list = soup.select(total_tag)
        
        score_list = []
        push_list = []

        for html in list :
            txts = []
            id = html['id']
            txts.append(id)
            for tag in score_tags :
                txt = html.select_one(tag)
                txt = txt.get_text().replace("\xa0","")
                txts.append(txt)
            score_list.append(txts)

            txts = ""
            first = True
            for tag in push_tags :
                txt = html.select_one(tag)
                if txt is None :
                    txt = ""
                else :
                    txt = txt.get_text().replace("\xa0","").replace("(","").replace(")","")

                if first :
                    txts = txt
                    first = False
                else :
                    txts = txts + "," + txt
            push_list.append(txts)

        return score_list, push_list

    def make_web_game_list_baseball (self,soup, total_tag, score_tags, push_tags) :
        list = soup.select(total_tag)
        
        score_list = []
        push_list = []

        for html in list :
            txts = []
            id = html['id']
            txts.append(id)
            for tag in score_tags :
                txt = html.select_one(tag)
                txt = txt.get_text().replace("\xa0","")
                txts.append(txt)
            score_list.append(txts)
        
        for game in score_list :
         
            web_id = game[WEBElem.WEBID]

            detail_url = self.detail_url.replace('id',web_id[4:])
            print("detail url :  " +detail_url)
            self.driver.get(detail_url)
            page = self.driver.page_source
            soup = BeautifulSoup(page, features = "lxml")

            txts = ""
            first = True
            
            for tag in push_tags :
                txt = soup.select_one(tag)
                if txt is None :
                    if (first) :
                        for i in range(0,10) :

                            self.driver.get(detail_url)
                            page = self.driver.page_source
                            soup = BeautifulSoup(page, features = "lxml")
                            txt = soup.select_one(tag)

                            if txt is not None :
                                txt = txt.get_text()
                                break
                            else :
                                txt = ""
                    else :
                        txt = ""
                else :
                    txt = txt.get_text().replace("\xa0","")

                if first :
                    txts = txt
                    first = False
                else :
                    txts = txts + "," + txt
          
            push_list.append(txts)

        return score_list, push_list

    def crawleAndParse (self) :
        self.set_options(self.prefs, self.arguments)
        self.load_webdriver()

        first = True

        for league, info in self.crawle_info.items() :

            url = info[CRAWLEINFO.URL]
            total_tag = info[CRAWLEINFO.TOTALTAG]
            score_tags = info[CRAWLEINFO.SCORETAG]
            push_tags = info[CRAWLEINFO.PUSHTAG]
            
            self.driver.get(url)

            if first is True :
                #self.set_gmt_zero()
                first = False

            #from this, parsing part
            page = self.driver.page_source
            soup = BeautifulSoup(page, features = "lxml")

            if league == 'KBO' or league == 'MLB' : 
                score_list,push_list = self.make_web_game_list_baseball(soup,total_tag,score_tags,push_tags)
            else : 
                score_list, push_list = self.make_web_game_list(soup,total_tag, score_tags, push_tags)

            game_list = []

            for i,score in enumerate(score_list) :
                push = push_list[i]
                score.append(push)
                game_list.append(score)

            self.web_results[league] = game_list
            for league,games in self.web_results.items() :
                print(league)
                if len(games) == 0 :
                    print('경기가 없습니다.')
                else :
                    print(games[0])

        self.quit_webdriver()

    def save_crawling_datas(self,path) :
        wb = openpyxl.Workbook()

        firstSheet = True
        for league, games in self.web_results.items() :

            if firstSheet :
                wb.active.title = league
                firstSheet = False
            else :
                wb.create_sheet(league)
            
            curSheet = wb[league]

            for row,game in enumerate(games) :
                for col,elem in enumerate(game) :
                    curSheet.cell(row+1,col+1).value = elem

        
        today = date.today().strftime("%Y-%m-%d")
        file_name = path + "/" + today + ".xlsx"
        wb.save(file_name)


def crawling () :
    print(datetime.now())
    config_path = './Configuration/config.ini'    
    sm = SportsWebCrawler(SeleniumCrawler.Driver.CHROME,config_path)
    sm.crawleAndParse()
    sm.save_crawling_datas('./CrawlingDatas')


"""
sched = BlockingScheduler()
sched.add_job(crawling, 'cron', hour = '00',minute="30", id="schdeuler_test")
sched.start()
"""

crawling()


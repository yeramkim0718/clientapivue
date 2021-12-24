from openpyxl import load_workbook
from time import gmtime, strftime
from Elem import*
import numpy as np

class SportsScoreConverter :

    team_matching_path = "./Configuration/team_matching.xlsx"
    team_dics = None

    # dic : key(team name in crawling site) value(team name in db)
    @staticmethod
    def make_team_matching_dic (sheet) :
        dic = {}
        for row in sheet.iter_rows(min_row = 2, max_col =3, min_col=1) :
            if row[0].value is None :
                break
            
            if row[1].value is not None :
                key = row[1].value
                value = row[2].value
                dic[key] = value

        return dic

    # dics : key(league name), value (dictionary of teamnames)
    @staticmethod
    def make_all_team_matching_dics() :
        workbook = load_workbook(SportsScoreConverter.team_matching_path)
        dics = {}
        for sheet in workbook :
            dic = SportsScoreConverter.make_team_matching_dic(sheet)
            dics[sheet.title] = dic

        return dics

    def __init__(self) :
        if SportsScoreConverter.team_dics is None :
            SportsScoreConverter.team_dics = SportsScoreConverter.make_all_team_matching_dics()
            
    def change_web_time_format(self,times) :

        n_times = []
        if str(type(times)) == "<class 'str'>" :
            time = times
            times = []
            times.append(time)

        for time in times :
            splited = time.replace(" ","").replace(":","").split(".")
            splited[2] = splited[2][0:4]   
            n_time = splited[1]+splited[0]+splited[2]

            year = strftime("%Y",gmtime())
            now = strftime("%m%d%H%M", gmtime()) 
            if(n_time > now) :
                n_time = str(int(year)-1) + n_time
            else :
                n_time = year + n_time
            n_times.append(n_time)
        
        if len(n_times) == 1 :
            return n_times[0]
        return n_times
    
    def change_web_team_format(self,league,team_names) :

        n_teams = []

        dics = SportsScoreConverter.team_dics
        for team in team_names : 
            if dics.get(league) is not None and dics.get(league).get(team):
                n_name = dics.get(league).get(team)
            else :
                n_name = team
            n_teams.append(n_name)
        
        return n_teams

    def change_web_score_format(self,league,scores) :

        n_scores = scores
        if league == 'IPL' :
            n_scores = []
            for score in scores :
                splited = score.split('/')
                n_score = splited[0]
                n_scores.append(n_score)

        return n_scores

    def change_web_format(self, web_results) :

        n_web_results = {}
        for league in web_results :
            lists = web_results.get(league)

            times = lists[WEBElem.TIME]
            homes = lists[WEBElem.HOME]
            aways = lists[WEBElem.AWAY]
            h_scores = lists[WEBElem.HOME_SCORE]
            a_scores = lists[WEBElem.AWAY_SCORE]

            lists[WEBElem.TIME] = self.change_web_time_format(times)
            lists[WEBElem.HOME] = self.change_web_team_format(league,homes)
            lists[WEBElem.AWAY] = self.change_web_team_format(league,aways)
            lists[WEBElem.HOME_SCORE] = self.change_web_score_format(league,h_scores)
            lists[WEBElem.AWAY_SCORE] = self.change_web_score_format(league,a_scores)

            n_web_results[league] = np.array(lists).T.tolist()

        return n_web_results

    def change_db_score_foramt(self, league, scores) :
        if league == 'IPL' :
            for i,score in enumerate(scores) :
                splited = score.split('/')
                score = splited[0]
                scores[i] = score
        
        return scores


    def change_db_sport_id(self,sport_ids) :
        mapper = { "GN2QKXNCTVXP8EV" : "BASKETBALL", "GNFB17MNF4TH4KM" : "FOOTBALL", "GN38CMNHHJM1Z26" : "AMERICAN FOOTBALL", "GN3FAZE2HJTVDJ9" : "BASEBALL", "GN7XZG4918F8AT5" : "ICEHOCKEY" , "GN48E1F0X7AYM76" :"CRICKET", "GNBF6D927MH62TN" : "GOLF", "GN8YS0R6P2XTHWQ" : "RUGBY LEAGUE", "GNDKH0FR9PFRYPR" : "AUSTRALIAN FOOTBALL"}
        sport_names = []
        for sport_id in sport_ids :
            sport_name = ""
            if sport_id in mapper.keys() :
                sport_name = mapper.get(sport_id)
            sport_names.append(sport_name)
        return sport_names

    def change_db_format(self, db_results) :
            
        n_db_results = {}

        for league in db_results :
            
            games = db_results.get(league)

            #convert 2nd tuple to 2nd list
            n_games = []
            for game in games :
                game = list(game)
                n_games.append(game)
            
            h_scores = list(list(zip(*n_games))[DBElem.HOME_SCORE])
            a_scores = list(list(zip(*n_games))[DBElem.AWAY_SCORE])
            sport_ids = list(list(zip(*n_games))[DBElem.SPORT])

            h_scores = self.change_db_score_foramt(league,h_scores)
            a_scores = self.change_db_score_foramt(league,a_scores)
            sport_names = self.change_db_sport_id(sport_ids)

            for i, score in enumerate(h_scores) :
                n_games[i][DBElem.HOME_SCORE] = h_scores[i]
                n_games[i][DBElem.AWAY_SCORE] = a_scores[i]
            
            n_db_results[league] = n_games
        
        return n_db_results

